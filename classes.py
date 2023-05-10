from openpyxl import load_workbook
from dataclasses import dataclass
from mapping import MEMBER_ID, MEMBER_FIRST, MEMBER_LAST, \
    SECTION, PRONOUNS, ROLE, \
    ADDRESS, CITY, STATE, ZIP, PHONE, EMAIL, STATUS

# Defines a Member data class. Each member in the spreadsheet will be
# an instance of the Member class, with each column of information
# mapping to an attribute of the class.

@dataclass
class Member:
    id: str
    first_name: str
    last_name: str
    pronouns: str
    section: str
    role: str
    email: str
    address: str
    city: str
    state: str
    zip: str
    phone: str
    status: str

def create_active_members(file):
    """Creates a list using the information from the spreadsheet
    of all active members. Returns a list of each member on the spreadsheet as 
    an instance of the Member dataclass, which includes all of the information
    from the columns in the spreadsheet."""
    workbook = load_workbook(filename=file)
    active_sheet = workbook["Active Members"]
    active_members = []
    for row in active_sheet.iter_rows(min_row=2, values_only=True):
        active_member = Member(id=row[MEMBER_ID],
                               first_name=row[MEMBER_FIRST],
                               last_name=row[MEMBER_LAST],
                               pronouns=row[PRONOUNS],
                               section=row[SECTION],
                               role=row[ROLE],
                               email=row[EMAIL],
                               address=row[ADDRESS],
                               city=row[CITY],
                               state=row[STATE],
                               zip=row[ZIP],
                               phone=row[PHONE],
                               status=row[STATUS])
        active_members.append(active_member)
    return active_members

def create_inactive_members(file):
    """Creates a list using the information from the spreadsheet
    of all inactive members. Returns a list of each member on the spreadsheet as 
    an instance of the Member dataclass, which includes all of the information
    from the columns in the spreadsheet."""
    workbook = load_workbook(filename=file)
    inactive_sheet = workbook["Inactive Members"]
    inactive_members = []
    for row in inactive_sheet.iter_rows(min_row=2, values_only=True):
        inactive_member = Member(id=row[MEMBER_ID],
                               first_name=row[MEMBER_FIRST],
                               last_name=row[MEMBER_LAST],
                               pronouns=row[PRONOUNS],
                               section=row[SECTION],
                               role=row[ROLE],
                               email=row[EMAIL],
                               address=row[ADDRESS],
                               city=row[CITY],
                               state=row[STATE],
                               zip=row[ZIP],
                               phone=row[PHONE],
                               status=row[STATUS])
        inactive_members.append(inactive_member)
    return inactive_members