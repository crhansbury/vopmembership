from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
import sys
from querymember import query_member_object
from emails import send_email, generate_email

def update_member(file, email):
    """Updates the member information in the spreadsheet. Prompts user for:
     name, pronouns, voice part, role, email address, phone number, or 
      mailing address."""
    # Open the file and iterate over the emails field to find the desired row
    workbook = load_workbook(file)
    sheet = workbook["Active Members"]
    edit_row = 0
    best_email = email
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, min_col=1, 
                                                  max_col=12, values_only=True), 
                                                  start=2):
        if email == row[11]:
            edit_row = row_num
            member_first = row[1]
            member_last = row[0]
    if edit_row == 0:
        print(f"{email} not found in {file}.")
    else:
        check = input(f"Edit information for {member_first} {member_last}? [y/n]\n")
        if check.lower().strip() == "y":
            while True:
                print("Please select which field you would like to edit\n",
                      "[1] First Name\n",
                      "[2] Last Name\n",
                      "[3] Pronouns\n",
                      "[4] Section\n",
                      "[5] Email\n",
                      "[6] Phone number\n",
                      "[7] Address\n",
                      "[8] Role")
                response = input("(enter a number from the menu above)\n")
                if response == "1":
                    first_name = input("Please enter new first name: ").title()
                    sheet.cell(row=edit_row, column=2, value=first_name)
                    member_first = first_name
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "2":
                    last_name = input("Please enter new last name: ").title()
                    sheet.cell(row=edit_row, column=1, value=last_name)
                    member_last = last_name
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "3":
                    pronouns = input("Please enter new pronouns: ").lower()
                    sheet.cell(row=edit_row, column=3, value=pronouns)
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "4":
                    section = input("Please enter new section: ").upper()
                    sheet.cell(row=edit_row, column=4, value=section)
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "5":
                    new_email = input("Please enter new email: ")
                    sheet.cell(row=edit_row, column=12, value=new_email)
                    best_email = new_email
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "6":
                    phone = input("Please enter new phone number: ")
                    sheet.cell(row=edit_row, column=11, value=phone)
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "7":
                    street_address = input("Please enter new street address: ").title()
                    sheet.cell(row=edit_row, column=7, value=street_address)
                    city = input("Please enter new city: ").title()
                    sheet.cell(row=edit_row, column=8, value=city)
                    state = input("Please enter new state: ").upper()
                    sheet.cell(row=edit_row, column=9, value=state)
                    zipcode = input("Please enter new zip code: ")
                    sheet.cell(row=edit_row, column=10, value=zipcode)
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                if response == "8":
                    role = input("Please enter new role: ").title()
                    sheet.cell(row=edit_row, column=6, value=role)
                    # Prompt user for another edit
                    again = input("Edit another field? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        break
                else:
                    print("Please enter a valid response from the menu.")
                    continue
            # Adding the date modified
            date = datetime.now()
            sheet["N2"] = "{} at {}".format(date.strftime("%m/%d/%Y"),
                                            date.strftime("%H:%M"))
            workbook.save(file)
            print(f"{member_first} {member_last} successfully edited.")
            subject, body = generate_email("email-templates/updated_member_template.txt",
                                           best_email, file)
            send_email(best_email, subject, body)


def inactive_member(file, attribute, attr_value):
    """Moves a member from the Active Member sheet to the Inactive Member 
    sheet."""
    member = query_member_object(file, attribute, attr_value)[0]
    # Open the inactive sheet and append the member to the inactive sheet
    check = input(f"Remove {member.first_name} {member.last_name} from Active Members? [y/n]\n")
    if check == "y":
        workbook = load_workbook(file)
        inactive_sheet = workbook["Inactive Members"]
        active_sheet = workbook["Active Members"]
        # Find the row that matches the email attribute
        active_sheet_row = 0
        for row_num, row in enumerate(active_sheet.iter_rows(min_row=2, min_col=12, 
                                                            max_col=12, 
                                                            values_only=True), 
                                                            start=2):
            if member.email == row[0]:
                active_sheet_row = row_num
        # Find the last row of inactive sheet
        inactive_sheet_row = inactive_sheet.max_row + 1
        # Copy/pasting all values from active sheet to inactive sheet
        for col_num, cell in enumerate(active_sheet[active_sheet_row], 1):
            value = cell.value
            inactive_sheet.cell(row=inactive_sheet_row, column=col_num, value=value)
        # Changing value for "status" column to "Inactive"
        inactive_sheet.cell(row=inactive_sheet_row, column=13, value="Inactive")
        # Formatting the cells
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
        for cell in inactive_sheet[inactive_sheet_row]:
            cell.border = thin_border
            cell.alignment = alignment
        # Delete the member from active member sheet
        active_sheet.delete_rows(idx=active_sheet_row)
        # Add date modified
        date = datetime.now()
        active_sheet["N2"] = "{} at {}".format(date.strftime("%m/%d/%Y"),
                                               date.strftime("%H:%M"))
        workbook.save(file)
        print("{} successfully moved to Inactive Members.".format(member.first_name))
        # Send an automatic email when this completes
        subject, body = generate_email("email-templates/inactive_member_template.txt",
                                    member.email, file)
        send_email(member.email, subject, body)
    else:
        print("Action cancelled.")

def active_member(file, attribute, attr_value):
    """Moves a member from the Inactive Member sheet to the Active Member 
    sheet. Takes the argument of a member attribute and value, and looks
    up that member. Copies the member over from the Inactive sheet to the 
    Active sheet."""
    member = query_member_object(file, attribute, attr_value)[0]
    # Open the active sheet and append the member to the active sheet
    check = input(f"Reinstate {member.first_name} {member.last_name} to Active Members? [y/n]\n")
    if check == "y":
        workbook = load_workbook(file)
        inactive_sheet = workbook["Inactive Members"]
        active_sheet = workbook["Active Members"]
        # Find the row that matches the email attribute
        inactive_sheet_row = 0
        for row_num, row in enumerate(inactive_sheet.iter_rows(min_row=2, 
                                                               min_col=12, 
                                                               max_col=12, 
                                                               values_only=True), 
                                                               start=2):
            if member.email == row[0]:
                inactive_sheet_row = row_num
        # Find the last row of inactive sheet
        active_sheet_row = active_sheet.max_row + 1
        # Copy/pasting all values from inactive sheet to active sheet
        for col_num, cell in enumerate(inactive_sheet[inactive_sheet_row], 1):
            value = cell.value
            active_sheet.cell(row=active_sheet_row, column=col_num, value=value)
        # Changing "status" col value to "Active"
        active_sheet.cell(row=active_sheet_row, column=13, value="Active")        
        # Formatting the cells
        thin_border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
        for cell in active_sheet[active_sheet_row]:
            cell.border = thin_border
            cell.alignment = alignment
        # Delete member from inactive member sheet
        inactive_sheet.delete_rows(idx=inactive_sheet_row)
        # Add date modified
        date = datetime.now()
        active_sheet["N2"] = "{} at {}".format(date.strftime("%m/%d/%Y"),
                                        date.strftime("%H:%M"))
        workbook.save(file)
        print("{} successfully moved to Active Members.".format(member.first_name))
        # Send an automatic email when this completes
        subject, body = generate_email("email-templates/active_member_template.txt",
                                    member.email, file)
        send_email(member.email, subject, body)
    else:
        print("Action cancelled.")