from openpyxl import load_workbook
from classes import Member
from mapping import MEMBER_ID, MEMBER_FIRST, MEMBER_LAST, \
    SECTION, PRONOUNS, ROLE, \
    ADDRESS, CITY, STATE, ZIP, PHONE, EMAIL 

def create_active_dictionary(file):
    """Creates a dictionary using the information from the spreadsheet
    of all active members. The key is the member id, and the value is 
    an instance of the Member dataclass, which includes all of the information
    from the columns in the spreadsheet."""
    workbook = load_workbook(filename=file)
    active_sheet = workbook["Active Members"]
    active_members = {}
    for row in active_sheet.iter_rows(min_row=2, values_only=True):
        active_member = Member(id=row[MEMBER_ID],
                               first_name=row[MEMBER_FIRST],
                               last_name=row[MEMBER_LAST],
                               pronouns=row[PRONOUNS],
                               section=row[SECTION],
                               role=row[ROLE],
                               email=row[EMAIL],
                               address=row[ADDRESS],
                               city=row[CITY],
                               state=row[STATE],
                               zip=row[ZIP],
                               phone=row[PHONE])
        active_members[active_member.id] = active_member
    return active_members

def create_inactive_dictionary(file):
    """Creates a dictionary using the information from the spreadsheet
    of all inactive members."""
    workbook = load_workbook(filename=file)
    inactive_sheet = workbook["Inactive Members"]
    inactive_members = {}
    for row in inactive_sheet.iter_rows(min_row=2, values_only=True):
        inactive_member = Member(id=row[MEMBER_ID],
                               first_name=row[MEMBER_FIRST],
                               last_name=row[MEMBER_LAST],
                               pronouns=row[PRONOUNS],
                               section=row[SECTION],
                               role=row[ROLE],
                               email=row[EMAIL],
                               address=row[ADDRESS],
                               city=row[CITY],
                               state=row[STATE],
                               zip=row[ZIP],
                               phone=row[PHONE])
        inactive_members[inactive_member.id] = inactive_member
    return inactive_members

def main():
    """Prompts user for input. Asks if user would like to enter a new 
    member into the spreadsheet, change member information, 
    make a member inactive, or query the spreadsheet for information"""
    pass

inactive_dict = create_inactive_dictionary("vopmembership_data.xlsx")
print(inactive_dict)