from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
import sys
import os
from emails import generate_email, send_email
from newmember import create_member_id
from querymember import query_active_member, query_member_object
from updatemember import update_member, inactive_member, active_member


def add_member(file):
    """Adds a member to the file. Takes criteria for:
    name, pronouns, voice part, role, email address, phone number, and
    mailing address. Appends file with this information. Generates a unique
    member ID and assigns it to the new member. Enters status as 'Active'
    automatically. Updates a field that indicates the date of the latest update."""
    # Prompt the user for info
    print("Please enter the information for the new member.")
    first_name = input("First Name: ").title()
    last_name = input("Last Name: ").title()
    pronouns = input("Pronouns: ").lower()
    section = input("Section: ").upper()
    role = input("Role: ").title()
    email = input("Email: ")
    phone = input("Phone Number: ")
    address = input("Street Address: ").title()
    city = input("City: ").title()
    state = input("State: ").upper()
    zipcode = input("Zip Code: ")
    status = "Active"
    # Generate new member ID
    member_id = create_member_id(file)
    # Open the workbook and load the "Active Member" sheet and append info
    workbook = load_workbook(filename=file)
    sheet = workbook["Active Members"]
    row = (
        last_name,
        first_name,
        pronouns,
        section,
        member_id,
        role,
        address,
        city,
        state,
        zipcode,
        phone,
        email,
        status,
    )
    sheet.append(row)
    # Formatting the new row to match the rest of the file
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    new_row = sheet.max_row
    for cell in sheet[new_row]:
        cell.border = thin_border
        cell.alignment = alignment
    # Adding the date modified
    date = datetime.now()
    sheet["N2"] = "{} at {}".format(date.strftime("%m/%d/%Y"), date.strftime("%H:%M"))
    workbook.save(file)
    workbook.close()
    print("✅ {} {} successfully added to {}.".format(first_name, last_name, file))
    # Generating and sending welcome email once new member has been successfully
    # appended to file
    try:
        subject, body = generate_email(
            "email-templates/new_member_template.txt", email, file
        )
        send_email(email, subject, body)
    except:
        pass


def all_active_email(template, file):
    """Sends an email to every member in the Active Members sheet of the
    spreadsheet passed into the 'file' parameter. Uses the email
    template passed into the 'template' parameter."""
    # Make a lsit of all emails from the Active Members sheet
    active_emails = query_active_member(file, "email")
    # Send an email to each email in list
    for email in active_emails:
        subject, body = generate_email(template, email, file)
        send_email(email, subject, body)


def search_member(file, attribute, attr_value):
    """Searches the file for members that match the search criteria.
    Prints all the information for the members who match the search.
    Numbers all of the results. Uses the query_member_object() function to
    search the file."""
    # Finding all members who match search criteria
    query_list = query_member_object(file, attribute, attr_value)
    if query_list == None:
        print("❌ No members meet your search criteria.")
    else:
        print(f"{len(query_list)} member(s) meet your search criteria.")
        member_count = 1
        # Print out information for each member
        for member in query_list:
            print("⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽⎽")
            print(f"Result number {member_count}:")
            print(
                f" Name: {member.first_name} {member.last_name}\n",
                f"Pronouns: {member.pronouns}\n",
                f"Section: {member.section}\n",
                f"Role: {member.role}\n",
                f"Member ID: {member.id}\n",
                f"Email: {member.email}\n",
                f"Phone number: {member.phone}\n",
                f"Address: {member.address}, {member.city} {member.state} {member.zip}\n",
                f"Status: {member.status}",
            )
            member_count += 1  # Incrementing member count


def main(file):
    """Prints out a main menu, asking if user would like to enter a new member
    into the file, change member information, make a member active/inactive,
    query the file for information, or send an email to all active members.
    Performs these actions accordingly."""
    # Test the file to ensure it is in the right format.
    if os.path.exists(file):
        workbook = load_workbook(file)
        try:
            active_sheet = workbook["Active Members"]
            try:
                inactive_sheet = workbook["Inactive Members"]
            except KeyError:
                print(
                    f"❌ {file} is not formatted correctly. Please ensure "
                    "file includes a sheet titled 'Inactive Members'."
                )
                sys.exit(2)
            column_names = [
                "last name",
                "first name",
                "pronouns",
                "section",
                "member id",
                "role",
                "address",
                "city",
                "state",
                "zip",
                "phone",
                "email",
                "status",
                "date modified",
            ]
            for col_num, cell in enumerate(active_sheet[1]):
                if cell.value != column_names[col_num]:
                    print(
                        f"❌ {file} is not formatted correctly. Please ensure the"
                        " file has 14 columns, named:"
                    )
                    print(*column_names, sep="\n")
                    sys.exit(3)
        except KeyError:
            print(
                f"❌ {file} is not formatted correctly. Please ensure file \
                  includes a sheet titled 'Active Members'."
            )
            sys.exit(4)
        workbook.close()
    else:
        print(f"❌ {file} does not exist.")
        sys.exit(1)
    while True:
        print("⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥")
        print(
            "   Welcome to the VOP Membership Portal!\n" "   What would you like to do?"
        )
        print(
            "🟥 [1] Add new members\n"
            "🟧 [2] Remove members from Active Members\n"
            "🟨 [3] Reinstate members to Active Members\n"
            "🟩 [4] Update active members\n"
            "🟦 [5] Search for member information\n"
            "🟪 [6] Send an email to all active members\n"
            "⬛️ [7] Exit"
        )
        print("⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥⪥")
        response = input("Please enter a number from the main menu:\n")
        if response == "7":
            print("👋 Goodbye!")
            break
        elif response == "1":
            # Double check that the user made the right choice
            check = input(
                "You have selected '🟥 Add new members'. Proceed?\n"
                "(enter 'y' to continue, or 'n' to go back to main menu):\n"
            )
            if check.lower() == "y":
                while True:
                    add_member(file)
                    # Check if the user wishes to add another member, or go back
                    # to main menu
                    again = input("Would you like to add another member? [y/n]\n")
                    if again.lower() == "y":
                        continue
                    else:
                        break
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        elif response == "2":
            check = input(
                "You have selected '🟧 Remove members from Active "
                "Members'. Proceed? [y/n]\n"
            )
            if check.lower().strip() == "y":
                while True:
                    while True:
                        print(
                            "Please choose search criteria for the member "
                            "you wish to remove:\n"
                            "[1] Email\n"
                            "[2] Member ID"
                        )
                        response = input("")
                        attribute = ""
                        if response == "1":
                            attribute = "email"
                            break
                        elif response == "2":
                            attribute = "id"
                            break
                        else:
                            print("❌ Please select a valid item from the menu.")
                            continue
                    attr_value = input(f"Please enter the member's {attribute}:\n")
                    try:
                        inactive_member(file, attribute, attr_value)
                    except TypeError:
                        continue
                    again = input("Would you like to remove another member? [y/n]\n")
                    if again.lower().strip() == "y":
                        continue
                    else:
                        print("⬅️  Returning to Main Menu.")
                        break
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        elif response == "3":
            check = input(
                "You have selected '🟨 Reinstate members to Active "
                "Members'. Proceed? [y/n]\n"
            )
            if check.lower().strip() == "y":
                while True:
                    while True:
                        print(
                            "Please choose search criteria for the member "
                            "you wish to reinstate:\n"
                            "[1] Email\n"
                            "[2] Member ID"
                        )
                        response = input("")
                        attribute = ""
                        if response == "1":
                            attribute = "email"
                            break
                        elif response == "2":
                            attribute = "id"
                            break
                        else:
                            print("❌ Please select a valid item from the menu.")
                            continue
                    attr_value = input(f"Please enter the member's {attribute}:\n")
                    try:
                        active_member(file, attribute, attr_value)
                    except TypeError:
                        continue
                    again = input("Would you like to reinstate another member? [y/n]\n")
                    if again.lower().strip() == "y":
                        continue
                    else:
                        print("⬅️  Returning to Main Menu.")
                        break
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        elif response == "4":
            check = input(
                "You have selected '🟩 Update active members'. " "Proceed? [y/n]\n"
            )
            if check.lower().strip() == "y":
                while True:
                    email = input(
                        "Please enter the email for the member whose "
                        "information you wish you update:\n"
                    )
                    update_member(file, email)
                    again = input("Update another member? [y/n]\n").lower().strip()
                    if again == "y":
                        continue
                    else:
                        print("⬅️  Returning to Main Menu.")
                        break
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        elif response == "5":
            check = input(
                "You have selected '🟦 Search for member information'. "
                "Proceed? [y/n]\n"
            )
            if check.lower().strip() == "y":
                while True:
                    print(
                        "Please choose the search criteria:\n",
                        "[1]  First Name\n",
                        "[2]  Last Name\n",
                        "[3]  Pronouns\n",
                        "[4]  Section\n",
                        "[5]  Role\n",
                        "[6]  Member ID\n",
                        "[7]  Email\n",
                        "[8]  Phone number\n",
                        "[9]  Street Address\n",
                        "[10] City\n",
                        "[11] State\n",
                        "[12] Zip Code\n",
                        "[13] Status",
                    )
                    reply = input("(choose a number from the menu above)\n")
                    attribute = ""
                    attr_value = ""
                    if reply == "1":
                        attribute = "first_name"
                        attr_value = input(
                            "Please enter the member's first name:\n"
                        ).title()
                    elif reply == "2":
                        attribute = "last_name"
                        attr_value = input(
                            "Please enter the member's last name:\n"
                        ).title()
                    elif reply == "3":
                        attribute = "pronouns"
                        attr_value = input(
                            "Please enter the member's pronouns:\n"
                        ).lower()
                    elif reply == "4":
                        attribute = "section"
                        attr_value = input(
                            "Please enter the member's section:\n"
                        ).upper()
                    elif reply == "5":
                        attribute = "role"
                        attr_value = input("Please enter the member's role:\n").title()
                    elif reply == "6":
                        attribute = "id"
                        attr_value = input("Please enter the member's member ID:\n")
                    elif reply == "7":
                        attribute = "email"
                        attr_value = input("Please enter the member's email:\n")
                    elif reply == "8":
                        attribute = "phone"
                        attr_value = input("Please enter the member's phone number:\n")
                    elif reply == "9":
                        attribute = "address"
                        attr_value = input(
                            "Please enter the member's street address:\n"
                        ).title()
                    elif reply == "10":
                        attribute = "city"
                        attr_value = input("Please enter the member's city:\n").title()
                    elif reply == "11":
                        attribute = "state"
                        attr_value = input("Please enter the member's state:\n").upper()
                    elif reply == "12":
                        attribute = "zip"
                        try:
                            attr_value = int(
                                input("Please enter the member's zip code:\n")
                            )
                        except ValueError:
                            print("❌ Not a valid zip code.")
                            continue
                    elif reply == "13":
                        attribute = "status"
                        attr_value = input(
                            "Please enter the member's status:\n"
                        ).title()
                    else:
                        print("❌ Please enter a valid number from the menu.")
                        continue
                    search_member(file, attribute, attr_value)
                    again = input("Would you like to perform another search? [y/n]\n")
                    if again.lower().strip() == "y":
                        continue
                    else:
                        print("⬅️  Returning to Main Menu.")
                        break
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        elif response == "6":
            check = input(
                "You have selected '🟪 Send an email to all active "
                "members'. Proceed? [y/n]\n"
            )
            if check == "y":
                template = input("Please enter the email template file: \n")
                all_active_email(template, file)
            else:
                print("⬅️  Returning to Main Menu.")
                continue
        else:
            print("❌ Not a valid response. Please enter a number from the main menu.")
            continue


if __name__ == "__main__":
    file = input("Enter the name of the spreadsheet:\n")
    if file == "":
        file = "data-files/vopmembership_data.xlsx"
    main(file)
